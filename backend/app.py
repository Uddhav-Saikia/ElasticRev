"""
ElasticRev - Dynamic Pricing Optimization API
Main Flask application with REST endpoints
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from models import db, Product, Sale, PriceHistory, ElasticityResult, Scenario, CompetitorPrice
from elasticity import ElasticityCalculator, calculate_revenue_optimization
from scenarios import ScenarioSimulator
from config import *
import pandas as pd
from datetime import datetime, timedelta
from sqlalchemy import func, and_, desc
import os
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import io

from flask import send_from_directory


# Initialize Flask app
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = SQLALCHEMY_DATABASE_URI
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = SQLALCHEMY_TRACK_MODIFICATIONS
app.config['SECRET_KEY'] = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# Initialize extensions
from flask import jsonify

db.init_app(app)
CORS(app, origins=CORS_ORIGINS)

# Initialize calculators
elasticity_calculator = ElasticityCalculator()
scenario_simulator = ScenarioSimulator()

# Database initialization flag
_db_initialized = False

def initialize_database():
    """Initialize database and seed data on first run"""
    global _db_initialized
    if _db_initialized:
        return
    
    with app.app_context():
        # Create all tables
        db.create_all()
        
        # Import seed function
        from database import seed_database_if_empty
        
        # Seed data if database is empty
        seed_database_if_empty(app)
        
        _db_initialized = True

# Serve React frontend static files from dist/
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    dist_dir = os.path.join(os.path.dirname(__file__), 'dist')
    file_path = os.path.join(dist_dir, path)
    if path != "" and os.path.exists(file_path):
        return send_from_directory(dist_dir, path)
    else:
        return send_from_directory(dist_dir, 'index.html')


# ==================== Error Handlers ====================

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Resource not found'}), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return jsonify({'error': 'Internal server error'}), 500


# ==================== Health Check ====================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Initialize database on first request
        initialize_database()
        
        # Check database connection
        product_count = Product.query.count()
        sales_count = Sale.query.count()
        
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'products': product_count,
            'sales': sales_count,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500


# ==================== Products API ====================

@app.route('/api/products', methods=['GET'])
def get_products():
    """Get all products with optional filtering"""
    try:
        category = request.args.get('category')
        search = request.args.get('search')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        query = Product.query
        
        if category:
            query = query.filter_by(category=category)
        
        if search:
            query = query.filter(Product.name.ilike(f'%{search}%'))
        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        products = [product.to_dict() for product in pagination.items]
        
        return jsonify({
            'products': products,
            'total': pagination.total,
            'page': page,
            'per_page': per_page,
            'pages': pagination.pages
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/products/<int:product_id>', methods=['GET'])
def get_product(product_id):
    """Get product details"""
    try:
        product = Product.query.get_or_404(product_id)
        
        # Get sales statistics
        sales_stats = db.session.query(
            func.count(Sale.id).label('total_sales'),
            func.sum(Sale.quantity).label('total_quantity'),
            func.sum(Sale.revenue).label('total_revenue'),
            func.sum(Sale.profit).label('total_profit'),
            func.avg(Sale.price).label('avg_price')
        ).filter_by(product_id=product_id).first()
        
        # Get latest elasticity
        latest_elasticity = ElasticityResult.query.filter_by(
            product_id=product_id
        ).order_by(ElasticityResult.calculation_date.desc()).first()
        
        result = product.to_dict()
        result['statistics'] = {
            'total_sales': sales_stats.total_sales or 0,
            'total_quantity': float(sales_stats.total_quantity or 0),
            'total_revenue': float(sales_stats.total_revenue or 0),
            'total_profit': float(sales_stats.total_profit or 0),
            'avg_price': float(sales_stats.avg_price or 0)
        }
        
        if latest_elasticity:
            result['elasticity'] = latest_elasticity.to_dict()
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/products/categories', methods=['GET'])
def get_categories():
    """Get all product categories"""
    try:
        categories = db.session.query(
            Product.category,
            func.count(Product.id).label('count')
        ).group_by(Product.category).all()
        
        return jsonify({
            'categories': [
                {'name': cat[0], 'count': cat[1]} 
                for cat in categories
            ]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Sales API ====================

@app.route('/api/sales', methods=['GET'])
def get_sales():
    """Get sales data with filtering"""
    try:
        product_id = request.args.get('product_id', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 100))
        
        query = Sale.query
        
        if product_id:
            query = query.filter_by(product_id=product_id)
        
        if start_date:
            query = query.filter(Sale.date >= datetime.fromisoformat(start_date).date())
        
        if end_date:
            query = query.filter(Sale.date <= datetime.fromisoformat(end_date).date())
        
        query = query.order_by(Sale.date.desc())
        
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        sales = [sale.to_dict() for sale in pagination.items]
        
        return jsonify({
            'sales': sales,
            'total': pagination.total,
            'page': page,
            'per_page': per_page
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/sales/summary', methods=['GET'])
def get_sales_summary():
    """Get sales summary statistics"""
    try:
        product_id = request.args.get('product_id', type=int)
        days = int(request.args.get('days', 30))
        
        date_threshold = datetime.now().date() - timedelta(days=days)
        
        query = Sale.query.filter(Sale.date >= date_threshold)
        
        if product_id:
            query = query.filter_by(product_id=product_id)
        
        summary = db.session.query(
            func.count(Sale.id).label('total_transactions'),
            func.sum(Sale.quantity).label('total_quantity'),
            func.sum(Sale.revenue).label('total_revenue'),
            func.sum(Sale.profit).label('total_profit'),
            func.avg(Sale.price).label('avg_price'),
            func.avg(Sale.quantity).label('avg_quantity')
        ).filter(Sale.date >= date_threshold)
        
        if product_id:
            summary = summary.filter_by(product_id=product_id)
        
        result = summary.first()
        
        return jsonify({
            'period_days': days,
            'total_transactions': result.total_transactions or 0,
            'total_quantity': float(result.total_quantity or 0),
            'total_revenue': float(result.total_revenue or 0),
            'total_profit': float(result.total_profit or 0),
            'avg_price': float(result.avg_price or 0),
            'avg_quantity': float(result.avg_quantity or 0),
            'avg_margin': round((result.total_profit / result.total_revenue * 100) if result.total_revenue else 0, 2)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Elasticity API ====================

@app.route('/api/elasticity/calculate', methods=['POST'])
def calculate_elasticity():
    """Calculate price elasticity for a product"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        model_type = data.get('model_type', 'linear_regression')
        
        if not product_id:
            return jsonify({'error': 'product_id is required'}), 400
        
        # Convert dates
        if start_date:
            start_date = datetime.fromisoformat(start_date).date()
        if end_date:
            end_date = datetime.fromisoformat(end_date).date()
        
        result = elasticity_calculator.calculate_elasticity(
            product_id, start_date, end_date, model_type
        )
        
        if 'error' in result:
            return jsonify(result), 400
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 400


@app.route('/api/elasticity/products/<int:product_id>', methods=['GET'])
def get_product_elasticity(product_id):
    """Get elasticity results for a product"""
    try:
        latest = request.args.get('latest', 'true').lower() == 'true'
        
        if latest:
            elasticity = ElasticityResult.query.filter_by(
                product_id=product_id
            ).order_by(ElasticityResult.calculation_date.desc()).first()
            
            if not elasticity:
                return jsonify({'error': 'No elasticity data found'}), 404
            
            return jsonify(elasticity.to_dict())
        else:
            elasticities = ElasticityResult.query.filter_by(
                product_id=product_id
            ).order_by(ElasticityResult.calculation_date.desc()).all()
            
            return jsonify({
                'elasticities': [e.to_dict() for e in elasticities]
            })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/elasticity/curve/<int:product_id>', methods=['GET'])
def get_elasticity_curve(product_id):
    """Get elasticity curve data for visualization"""
    try:
        curve_data = elasticity_calculator.get_elasticity_curve(product_id)
        
        if 'error' in curve_data:
            return jsonify(curve_data), 400
        
        return jsonify(curve_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/elasticity/bulk-calculate', methods=['POST'])
def bulk_calculate_elasticity():
    """Calculate elasticity for multiple products"""
    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])
        model_type = data.get('model_type', 'linear_regression')
        
        if not product_ids:
            # Calculate for all products
            products = Product.query.all()
            product_ids = [p.id for p in products]
        
        results = []
        errors = []
        
        for product_id in product_ids:
            try:
                result = elasticity_calculator.calculate_elasticity(
                    product_id, model_type=model_type
                )
                
                if 'error' not in result:
                    results.append({
                        'product_id': product_id,
                        'success': True,
                        'elasticity': result['elasticity_coefficient'],
                        'type': result['elasticity_type']
                    })
                else:
                    errors.append({
                        'product_id': product_id,
                        'error': result['error']
                    })
            except Exception as e:
                errors.append({
                    'product_id': product_id,
                    'error': str(e)
                })
        
        return jsonify({
            'success': results,
            'errors': errors,
            'total_calculated': len(results),
            'total_errors': len(errors)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Scenarios API ====================

@app.route('/api/scenarios/simulate', methods=['POST'])
def simulate_scenario():
    """Simulate a what-if pricing scenario"""
    try:
        data = request.get_json()
        product_id = data.get('product_id')
        new_price = data.get('new_price')
        simulation_days = data.get('simulation_days', 30)
        scenario_name = data.get('scenario_name')
        
        if not product_id or not new_price:
            return jsonify({'error': 'product_id and new_price are required'}), 400
        
        result = scenario_simulator.simulate_scenario(
            product_id, new_price, simulation_days, scenario_name
        )
        
        if 'error' in result:
            return jsonify(result), 400
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 400


@app.route('/api/scenarios', methods=['GET'])
def get_scenarios():
    """Get all scenarios"""
    try:
        product_id = request.args.get('product_id', type=int)
        limit = int(request.args.get('limit', 20))
        
        query = Scenario.query
        
        if product_id:
            query = query.filter_by(product_id=product_id)
        
        scenarios = query.order_by(Scenario.created_at.desc()).limit(limit).all()
        
        return jsonify({
            'scenarios': [s.to_dict() for s in scenarios]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/scenarios/<int:scenario_id>', methods=['GET'])
def get_scenario(scenario_id):
    """Get scenario details"""
    try:
        scenario = Scenario.query.get_or_404(scenario_id)
        return jsonify(scenario.to_dict())
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/scenarios/compare', methods=['POST'])
def compare_scenarios():
    """Compare multiple scenarios"""
    try:
        data = request.get_json()
        scenario_ids = data.get('scenario_ids', [])
        
        if not scenario_ids:
            return jsonify({'error': 'scenario_ids are required'}), 400
        
        result = scenario_simulator.compare_scenarios(scenario_ids)
        
        if 'error' in result:
            return jsonify(result), 400
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/scenarios/bulk-simulate', methods=['POST'])
def bulk_simulate_scenarios():
    """Simulate multiple scenarios"""
    try:
        data = request.get_json()
        product_ids = data.get('product_ids', [])
        price_changes = data.get('price_changes', [-10, -5, 5, 10])  # percentages
        
        if not product_ids:
            return jsonify({'error': 'product_ids are required'}), 400
        
        result = scenario_simulator.bulk_simulate(product_ids, price_changes)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Recommendations API ====================

@app.route('/api/recommendations', methods=['GET'])
def get_recommendations():
    """Get pricing recommendations for all products"""
    try:
        category = request.args.get('category')
        
        query = Product.query
        
        if category:
            query = query.filter_by(category=category)
        
        products = query.all()
        
        recommendations = []
        
        for product in products:
            latest_elasticity = ElasticityResult.query.filter_by(
                product_id=product.id
            ).order_by(ElasticityResult.calculation_date.desc()).first()
            
            if latest_elasticity:
                recommendations.append({
                    'product_id': product.id,
                    'product_name': product.name,
                    'category': product.category,
                    'current_price': product.current_price,
                    'optimal_price': latest_elasticity.optimal_price,
                    'recommended_action': latest_elasticity.recommended_action,
                    'expected_revenue_change': latest_elasticity.expected_revenue_change,
                    'elasticity_type': latest_elasticity.elasticity_type,
                    'elasticity_coefficient': latest_elasticity.elasticity_coefficient
                })
        
        # Sort by expected revenue change
        recommendations.sort(key=lambda x: x.get('expected_revenue_change', 0), reverse=True)
        
        return jsonify({
            'recommendations': recommendations,
            'total': len(recommendations)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/recommendations/<int:product_id>', methods=['GET'])
def get_product_recommendation(product_id):
    """Get pricing recommendation for specific product"""
    try:
        optimization = calculate_revenue_optimization(product_id)
        
        if 'error' in optimization:
            return jsonify(optimization), 400
        
        return jsonify(optimization)
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Analytics API ====================

@app.route('/api/analytics/dashboard', methods=['GET'])
def get_dashboard_analytics():
    """Get dashboard KPIs and analytics"""
    try:
        days = int(request.args.get('days', 30))
        date_threshold = datetime.now().date() - timedelta(days=days)
        
        # Overall metrics
        overall = db.session.query(
            func.sum(Sale.revenue).label('total_revenue'),
            func.sum(Sale.profit).label('total_profit'),
            func.sum(Sale.quantity).label('total_quantity'),
            func.count(func.distinct(Sale.product_id)).label('products_sold')
        ).filter(Sale.date >= date_threshold).first()
        
        # Category performance
        category_perf = db.session.query(
            Product.category,
            func.sum(Sale.revenue).label('revenue'),
            func.sum(Sale.profit).label('profit')
        ).join(Sale).filter(Sale.date >= date_threshold).group_by(
            Product.category
        ).all()
        
        # Elasticity distribution
        elasticity_dist = db.session.query(
            ElasticityResult.elasticity_type,
            func.count(ElasticityResult.id).label('count')
        ).group_by(ElasticityResult.elasticity_type).all()
        
        # Top products by revenue
        top_products = db.session.query(
            Product.id,
            Product.name,
            func.sum(Sale.revenue).label('revenue')
        ).join(Sale).filter(Sale.date >= date_threshold).group_by(
            Product.id, Product.name
        ).order_by(desc('revenue')).limit(10).all()
        
        return jsonify({
            'period_days': days,
            'overall': {
                'total_revenue': float(overall.total_revenue or 0),
                'total_profit': float(overall.total_profit or 0),
                'total_quantity': float(overall.total_quantity or 0),
                'products_sold': overall.products_sold or 0,
                'avg_margin': round((overall.total_profit / overall.total_revenue * 100) if overall.total_revenue else 0, 2)
            },
            'by_category': [
                {
                    'category': cat.category,
                    'revenue': float(cat.revenue),
                    'profit': float(cat.profit)
                }
                for cat in category_perf
            ],
            'elasticity_distribution': [
                {
                    'type': dist.elasticity_type,
                    'count': dist.count
                }
                for dist in elasticity_dist
            ],
            'top_products': [
                {
                    'id': prod.id,
                    'name': prod.name,
                    'revenue': float(prod.revenue)
                }
                for prod in top_products
            ]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ==================== Excel Export API ====================

def get_dashboard_analytics_data(days=30):
    """Helper function to get dashboard analytics data for export"""
    date_threshold = datetime.now().date() - timedelta(days=days)
    
    # Overall metrics
    overall = db.session.query(
        func.sum(Sale.revenue).label('total_revenue'),
        func.sum(Sale.profit).label('total_profit'),
        func.count(func.distinct(Sale.product_id)).label('products_sold')
    ).filter(Sale.date >= date_threshold).first()
    
    # Total products and products with elasticity
    total_products = Product.query.count()
    products_with_elasticity = ElasticityResult.query.distinct(ElasticityResult.product_id).count()
    
    return {
        'total_revenue': float(overall.total_revenue or 0),
        'total_profit': float(overall.total_profit or 0),
        'products_sold': overall.products_sold or 0,
        'total_products': total_products,
        'products_with_elasticity': products_with_elasticity
    }

# ==================== Diagnostics ====================

@app.route('/api/diagnostics/data-quality', methods=['GET'])
def data_quality():
    """Check data quality for elasticity calculations - useful for debugging"""
    try:
        products = Product.query.all()
        results = []
        total_sales = 0
        products_can_calculate = 0
        
        for product in products:
            sales_count = Sale.query.filter_by(product_id=product.id).count()
            has_elasticity = ElasticityResult.query.filter_by(product_id=product.id).first() is not None
            can_calculate = sales_count >= 10
            
            if can_calculate:
                products_can_calculate += 1
            total_sales += sales_count
            
            results.append({
                'product_id': product.id,
                'product_name': product.name,
                'sales_count': sales_count,
                'can_calculate': can_calculate,
                'has_elasticity': has_elasticity
            })
        
        return jsonify({
            'status': 'ok',
            'total_products': len(results),
            'total_sales': total_sales,
            'products_with_sufficient_data': products_can_calculate,
            'products_with_elasticity': sum(1 for r in results if r['has_elasticity']),
            'details': results
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'trace': traceback.format_exc()
        }), 500

@app.route('/api/export/excel', methods=['GET'])
def export_to_excel():
    """Export comprehensive pricing strategy report to Excel"""
    try:
        days = int(request.args.get('days', 30))
        
        # Get dashboard data
        analytics = get_dashboard_analytics_data(days)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Summary'
        
        # Title
        ws['A1'] = 'ElasticRev - Pricing Strategy Report'
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(size=10, italic=True)
        
        # Summary metrics
        row = 4
        ws[f'A{row}'] = 'Key Metrics'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        metrics = [
            ('Total Revenue (30d)', f"${analytics['total_revenue']:,.2f}"),
            ('Total Profit (30d)', f"${analytics['total_profit']:,.2f}"),
            ('Products Analyzed', analytics['total_products']),
            ('Products with Elasticity', analytics['products_with_elasticity']),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Products with elasticity
        row += 2
        ws[f'A{row}'] = 'Products with Elasticity Analysis'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        elasticity_results = ElasticityResult.query.order_by(
            ElasticityResult.calculation_date.desc()
        ).limit(50).all()
        
        if elasticity_results:
            ws[f'A{row}'] = 'Product'
            ws[f'B{row}'] = 'Elasticity'
            ws[f'C{row}'] = 'Type'
            ws[f'D{row}'] = 'Optimal Price'
            
            for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell].font = Font(bold=True)
                ws[cell].fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            
            row += 1
            for result in elasticity_results:
                product = Product.query.get(result.product_id)
                ws[f'A{row}'] = product.name if product else 'Unknown'
                ws[f'B{row}'] = round(result.elasticity_coefficient, 3)
                ws[f'C{row}'] = result.elasticity_type
                ws[f'D{row}'] = f"${result.optimal_price:.2f}" if result.optimal_price else 'N/A'
                row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'elasticrev-report-{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f'Excel export error: {e}')
        traceback.print_exc()
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ==================== Main ====================

# Temporary endpoint to initialize database tables in production
@app.route('/api/init-db', methods=['POST'])
def init_db():
    with app.app_context():
        db.create_all()
    return jsonify({"status": "ok", "message": "Database tables created."})

if __name__ == '__main__':
    with app.app_context():
        initialize_database()
    
    print("=" * 60)
    print("🚀 ElasticRev - Dynamic Pricing Optimization API")
    print("=" * 60)
    print(f"📍 Server running on: http://localhost:5000")
    print(f"📚 API documentation: http://localhost:5000/api/health")
    print("=" * 60)
    
    app.run(debug=DEBUG, host='0.0.0.0', port=5000)
